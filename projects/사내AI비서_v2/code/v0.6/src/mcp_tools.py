"""MCP 도구 모듈.

LangChain @tool 데코레이터로 정의된 4개의 MCP 도구를 제공한다.
정형 데이터(직원, 연차, 매출)는 PostgreSQL에서 조회하며,
비정형 데이터(사내 문서)는 ChromaDB 벡터 검색을 사용한다.
ChromaDB가 없으면 data/docs/ 원본 문서를 파싱하여 자동 구축한다.

IPO 패턴:
  Input  - 각 도구별 파라미터 (emp_no, dept, start_date, end_date, query, k)
  Process - PostgreSQL 조회 (정형) 또는 ChromaDB 검색 (비정형)
  Output - 구조화된 딕셔너리 (JSON 직렬화 가능)
"""

from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Any, Optional

# LangChain 도구 데코레이터
try:
    from langchain_core.tools import tool
except ImportError:
    def tool(func):  # type: ignore
        """langchain_core 없을 때 무해한 패스스루 데코레이터."""
        return func

# ---------------------------------------------------------------------------
# 1. DB 연결 헬퍼
# ---------------------------------------------------------------------------

def _run_query(sql: str, params: tuple = ()) -> list[dict]:
    """PostgreSQL 쿼리를 실행하고 결과를 반환한다.

    연결 실패 시 빈 리스트를 반환한다.

    Args:
        sql: 실행할 SQL 문자열.
        params: 쿼리 바인딩 파라미터 튜플.

    Returns:
        딕셔너리 리스트 또는 빈 리스트 (연결 실패 시).
    """
    try:
        import psycopg2
        import psycopg2.extras

        db_url = os.getenv("DATABASE_URL", "")
        if not db_url:
            host = os.getenv("POSTGRES_HOST", "localhost")
            port = os.getenv("POSTGRES_PORT", "5432")
            db = os.getenv("POSTGRES_DB", "rag_db")
            user = os.getenv("POSTGRES_USER", "rag_user")
            password = os.getenv("POSTGRES_PASSWORD", "")
            db_url = f"postgresql://{user}:{password}@{host}:{port}/{db}"

        conn = psycopg2.connect(db_url, connect_timeout=3)  # ①
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)  # ②
        cur.execute(sql, params)  # ③
        rows = [dict(row) for row in cur.fetchall()]  # ④
        cur.close()
        conn.close()
        return rows
    except Exception:
        return []  # DB 없으면 빈 리스트


# ---------------------------------------------------------------------------
# 2. DB 초기화 헬퍼
# ---------------------------------------------------------------------------

_DB_ERROR_MSG = (
    "PostgreSQL에 연결할 수 없습니다. "
    "docker-compose up -d 로 DB를 시작한 후 다시 시도하십시오."
)


def _build_vectorstore() -> Any:
    """ChromaDB VectorStore를 생성한다. 없으면 data/docs/에서 자동 구축한다.

    v0.5와 동일한 파이프라인: docs → 파싱 → 청킹 → 임베딩 → ChromaDB 저장.

    Returns:
        chromadb.Collection 또는 None (구축 실패 시).
    """
    try:
        import chromadb
        from chromadb.utils.embedding_functions import SentenceTransformerEmbeddingFunction
    except ImportError:
        print("[경고] chromadb 또는 sentence-transformers 미설치. 키워드 검색으로 동작합니다.")
        return None

    chroma_path = os.getenv("CHROMA_PERSIST_DIR", "./data/chroma_db")
    collection_name = os.getenv("CHROMA_COLLECTION_NAME", "metacoding_documents")
    ef = SentenceTransformerEmbeddingFunction(model_name="jhgan/ko-sroberta-multitask")

    client = chromadb.PersistentClient(path=chroma_path)

    # 기존 ChromaDB 데이터 확인
    try:
        collection = client.get_collection(collection_name, embedding_function=ef)
        if collection.count() > 0:
            print(f"[INFO] 기존 ChromaDB 로드: {chroma_path} ({collection.count()}건)")
            return collection
    except Exception:
        pass

    # ChromaDB 없음 → data/docs/ 원본 문서에서 자동 구축
    print("[INFO] ChromaDB가 없습니다. data/docs/ 원본 문서에서 자동 구축합니다.")
    docs = _parse_and_chunk_docs()
    if not docs:
        print("[경고] data/docs/에 문서가 없습니다. 키워드 검색으로 동작합니다.")
        return None

    collection = client.get_or_create_collection(collection_name, embedding_function=ef)
    for i, doc in enumerate(docs):
        collection.add(
            ids=[f"doc_{i}"],
            documents=[doc["content"]],
            metadatas=[{"source": doc["source"], "page": doc.get("page", 1)}],
        )
    print(f"[INFO] ChromaDB 자동 구축 완료: {len(docs)}건 → {chroma_path}")
    return collection


def _parse_and_chunk_docs(chunk_size: int = 500, overlap: int = 100) -> list[dict]:
    """data/docs/의 원본 PDF/DOCX/XLSX를 파싱→청킹하여 딕셔너리 리스트로 반환한다.

    v0.5의 _parse_and_chunk_docs()와 동일한 파이프라인.

    Returns:
        {"content": str, "source": str, "page": int} 형태의 딕셔너리 리스트.
    """
    import pypdf
    from docx import Document as DocxDocument
    import openpyxl

    docs_dir = Path(__file__).parent.parent / "data" / "docs"
    if not docs_dir.exists():
        return []

    documents: list[dict] = []
    for file_path in sorted(docs_dir.rglob("*")):
        suffix = file_path.suffix.lower()
        source = file_path.stem
        texts: list[tuple[str, int]] = []

        if suffix == ".pdf":
            try:
                with open(file_path, "rb") as f:
                    reader = pypdf.PdfReader(f)
                    for page_num, page in enumerate(reader.pages, start=1):
                        text = (page.extract_text() or "").strip()
                        if text:
                            texts.append((text, page_num))
                print(f"  [INFO] PDF 파싱: {file_path.name} ({len(reader.pages)}페이지)")
            except Exception:
                continue
        elif suffix == ".docx":
            try:
                doc = DocxDocument(str(file_path))
                full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                if full_text:
                    texts.append((full_text, 1))
                print(f"  [INFO] DOCX 파싱: {file_path.name}")
            except Exception:
                continue
        elif suffix == ".xlsx":
            try:
                wb = openpyxl.load_workbook(str(file_path), data_only=True)
                for idx, name in enumerate(wb.sheetnames, start=1):
                    ws = wb[name]
                    rows = []
                    for row in ws.iter_rows():
                        cells = [str(c.value).strip() for c in row if c.value is not None]
                        if cells:
                            rows.append(" | ".join(cells))
                    if rows:
                        texts.append(("\n".join(rows), idx))
                print(f"  [INFO] XLSX 파싱: {file_path.name} ({len(wb.sheetnames)}시트)")
            except Exception:
                continue

        # 청킹: chunk_size 단위, overlap 오버랩
        for text, page_num in texts:
            step = chunk_size - overlap
            start = 0
            while start < len(text):
                chunk = text[start:start + chunk_size].strip()
                if chunk:
                    documents.append({"content": chunk, "source": source, "page": page_num})
                start += step

    return documents


# 모듈 레벨 벡터스토어 캐시 (한 번만 구축)
_VECTORSTORE_CACHE: Any = None


def _get_vectorstore() -> Any:
    """벡터스토어 싱글톤을 반환한다."""
    global _VECTORSTORE_CACHE
    if _VECTORSTORE_CACHE is None:
        _VECTORSTORE_CACHE = _build_vectorstore()
    return _VECTORSTORE_CACHE


# ---------------------------------------------------------------------------
# 3. MCP 도구 정의
# ---------------------------------------------------------------------------

@tool
def leave_balance(emp_no: str) -> dict:
    """직원의 연차 잔여 일수를 조회한다.

    직원 번호(E001 형식) 또는 이름으로 연차 정보를 반환한다.
    PostgreSQL 연결이 필요하다 (docker-compose up -d).

    Args:
        emp_no: 직원 번호 (예: "E001") 또는 이름 (예: "김민준").

    Returns:
        직원 연차 정보 딕셔너리. 미발견 시 error 키 포함.
    """
    # ① DB 조회 시도 (이름 또는 번호)
    if emp_no.startswith("E") and emp_no[1:].isdigit():
        rows = _run_query(
            """
            SELECT e.emp_no, e.name, e.department,
                   l.total_days, l.used_days,
                   (l.total_days - l.used_days) AS remaining_days
            FROM employees e
            JOIN leave_balance l ON e.emp_no = l.emp_no
            WHERE e.emp_no = %s
            """,
            (emp_no,),
        )
    else:
        rows = _run_query(
            """
            SELECT e.emp_no, e.name, e.department,
                   l.total_days, l.used_days,
                   (l.total_days - l.used_days) AS remaining_days
            FROM employees e
            JOIN leave_balance l ON e.emp_no = l.emp_no
            WHERE e.name LIKE %s
            """,
            (f"%{emp_no}%",),
        )

    # ② DB 결과가 있으면 반환
    if rows:
        return rows[0]

    # ③ DB 연결 실패 시 에러 반환
    return {"error": f"직원 '{emp_no}'을(를) 찾을 수 없습니다. {_DB_ERROR_MSG}"}


@tool
def sales_sum(dept: str = "", start_date: str = "", end_date: str = "") -> dict:
    """부서별 또는 전체 매출 합계를 조회한다.

    기간 필터와 부서 필터를 적용한 매출 통계를 반환한다.
    PostgreSQL 연결이 필요하다 (docker-compose up -d).

    Args:
        dept: 부서명 (예: "영업부"). 빈 문자열이면 전체 집계.
        start_date: 조회 시작일 (YYYY-MM-DD 형식, 기본값: 2024-11-01).
        end_date: 조회 종료일 (YYYY-MM-DD 형식, 기본값: 2024-12-31).

    Returns:
        total_amount, record_count, top5 등을 포함한 딕셔너리.
    """
    # ① 파라미터 기본값 처리
    start = start_date or "2024-11-01"  # ①
    end = end_date or "2024-12-31"

    # ② DB 조회 시도
    dept_filter = f"AND e.department LIKE '%{dept}%'" if dept else ""
    rows = _run_query(
        f"""
        SELECT e.department, e.name AS employee_name,
               SUM(s.amount) AS total_amount, COUNT(*) AS record_count
        FROM sales s
        JOIN employees e ON s.emp_no = e.emp_no
        WHERE s.sale_date BETWEEN %s AND %s {dept_filter}
        GROUP BY e.department, e.name
        ORDER BY total_amount DESC
        """,
        (start, end),
    )

    # ③ DB 결과 가공
    if rows:
        grand_total = sum(int(r.get("total_amount") or 0) for r in rows)
        return {
            "total_amount": grand_total,
            "record_count": len(rows),
            "dept_filter": dept or "전체",
            "period": f"{start} ~ {end}",
            "top5": rows[:5],
        }

    # ④ DB 연결 실패 시 에러 반환
    return {"error": _DB_ERROR_MSG, "dept_filter": dept or "전체", "period": f"{start} ~ {end}"}


@tool
def list_employees(dept: str = "") -> dict:
    """직원 목록을 조회한다.

    부서명 필터를 적용하여 직원 정보를 반환한다.
    PostgreSQL 연결이 필요하다 (docker-compose up -d).

    Args:
        dept: 부서명 필터 (예: "영업부"). 빈 문자열이면 전체 조회.

    Returns:
        employees 리스트와 count를 포함한 딕셔너리.
    """
    # ① DB 조회 시도
    if dept:
        rows = _run_query(
            "SELECT emp_no, name, department, position, hire_date FROM employees WHERE department LIKE %s ORDER BY name",
            (f"%{dept}%",),
        )
    else:
        rows = _run_query(
            "SELECT emp_no, name, department, position, hire_date FROM employees ORDER BY department, name",
            (),
        )

    # ② DB 결과 반환
    if rows:
        return {"employees": rows, "count": len(rows), "dept_filter": dept or "전체"}

    # ③ DB 연결 실패 시 에러 반환
    return {"error": _DB_ERROR_MSG, "employees": [], "count": 0, "dept_filter": dept or "전체"}


@tool
def search_documents(query: str, k: int = 3) -> dict:
    """사내 문서에서 관련 내용을 벡터 검색한다.

    ChromaDB 벡터 검색을 사용한다. ChromaDB가 없으면 data/docs/ 원본 문서를
    파싱하여 자동 구축한 후 검색한다 (v0.5와 동일한 파이프라인).

    Args:
        query: 검색 쿼리 (자연어 질문).
        k: 반환할 최대 문서 수 (기본값: 3).

    Returns:
        results 리스트와 total_found를 포함한 딕셔너리.
    """
    # ① ChromaDB 벡터 검색 (없으면 자동 구축)
    collection = _get_vectorstore()  # ①
    if collection is not None:
        try:
            results = collection.query(query_texts=[query], n_results=k)  # ②

            docs = []
            for i, doc in enumerate(results["documents"][0]):
                docs.append({
                    "content": doc,
                    "source": results["metadatas"][0][i].get("source", "unknown"),
                    "score": round(1 - results["distances"][0][i], 4),
                })
            return {"results": docs, "total_found": len(docs), "search_mode": "vector"}
        except Exception:
            pass

    # ② 벡터스토어 구축 실패 시 키워드 검색 폴백
    docs_dir = Path(__file__).parent.parent / "data" / "docs"
    parsed = _parse_and_chunk_docs(chunk_size=1000, overlap=0)
    if not parsed:
        return {"results": [], "total_found": 0, "search_mode": "keyword_fallback"}

    query_lower = query.lower()
    scored = []
    for doc in parsed:
        score = sum(1 for w in query_lower.split() if w in doc["content"].lower())
        if score > 0:
            scored.append((score, doc))
    scored.sort(key=lambda x: x[0], reverse=True)

    results = [
        {"content": d["content"][:300], "source": d["source"], "score": s}
        for s, d in scored[:k]
    ]
    return {"results": results, "total_found": len(results), "search_mode": "keyword_fallback"}


# ---------------------------------------------------------------------------
# 4. 도구 목록 (에이전트에 전달)
# ---------------------------------------------------------------------------

ALL_TOOLS = [leave_balance, sales_sum, list_employees, search_documents]
